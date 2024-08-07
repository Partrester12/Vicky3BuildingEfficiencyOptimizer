#!/usr/bin/env python
# coding: utf-8

# In[103]:


from scipy.optimize import linprog
from pulp import *
import pandas as pd
import numpy as np
import math
import mystic
from mystic.solvers import diffev2
from mystic.solvers import fmin
import mystic.symbolic as ms
from mystic.monitors import VerboseMonitor
from openpyxl import load_workbook
import customtkinter


# In[104]:
#Tape and glue to fix problems quickly! Note that this is a temporary fix in order to get this thing working asap
#This file has been transformed purely into a function used by GUI.py. Is it ugly? Yes. Does it go against good coding practices? Yes. Am I tired+in a hurry and want this to work? Also Yes.
def OptimizePrices(app, statusText):
    # IF YOU DON'T WANT TO EDIT ANYTHING, THEN JUST PRESS 'RUN' AT THE TOP BAR, SELECT 'RUN ALL CELLS', AND SCROLL TO THE BOTTOM FOR RESULTS!!

    #Some constants used throughout the code
    #Would highly suggest you to not edit these

    BASEPRICES = [100, 60, 30, 30, 50, 40, 30, 60, 50, 20, 30, 200, 20, 30, 30, 40, 20, 30, 40, 40, 40, 30, 60, 60, 30, 40, 50, 30, 70, 80, 40, 30, 40, 20, 70, 50, 30, 50, 50, 70, 40, 40, 30, 50]

    con_try = BASEPRICES

    lab_try = BASEPRICES

    MAXPRICES = (np.array(BASEPRICES)*1.75).astype(int).tolist()

    GOODSNAMES = ["Soft wood","Hard wood","Iron","Coal","Tools","Steel","Fertilizer","Dye","Glass","Lead","Oil","Rubber","Silk","Explosives","Sulfur","Clippers","Engines","Steamers","Automobiles","Coffee","Fine art","Fruit","Liqour","Luxury clothes","Luxury furniture","Meat","Opium","Porcelain","Radios","Sugar","Tea","Telephones","Tobacco","Wine","Electricity","Services","Transportation","Paper","Groceries","Grain","Furniture","Fish","Fabric","Clothes"]
    #Sorting alphabetically. Should be done by hand but the dev is lazy...
    GOODSNAMES.sort()

    #This is the one exception! Change this if you feel like results are taking FOREVER
    MAX_NUMBER_ITERATIONS = 10000


    # In[105]:


        
    #Import the Excel sheet containing the buildings/PMs used to produce each good
    #Remember to edit this name to match the name of the excel sheet you have on your PC. By default we look for 'BuildingSheet.xlsx'
    data = pd.read_excel("BuildingSheet.xlsx")

    #Create the data frame based only on the buildigns we want included in the spreadsheet
    df = data[data['Included'].astype(str).str.contains("1")]

    #print(df)

    #New testing with 'looping' solutions as new starting points

    try:
        data_con = pd.read_excel("OptimizedPrices.xlsx", "Optimized for construction")
        con_try = data_con['Optimal price per construction'].astype(float).values.tolist()

        data_lab = pd.read_excel("OptimizedPrices.xlsx", "Optimized for labor")
        lab_try = data_lab['Optimal price per labor'].astype(float).values.tolist()

    except:
        print("No existing file found, continuing with base prices as initial guesses in case linear optimization cannot be performed")


    # In[106]:


    #Construct the matrices needed for creating our linear minimizing problem (maths stuff)
    #Optimization for labor is done per 100 pops

    #Calculating the scalar factor for outputs and inputs when optimizing for construction
    #Also taking into account potential construction bonuses (aka companies)
    scalar_per_construction=df['Construction'].values*(1.0-df['ConBonus'].values)

    #print((1.0-df['ConBonus'].values))

    #print(scalar_per_construction)

    #Least common multiplier to make everything into integers
    con_lcm = np.lcm.reduce(scalar_per_construction.astype(np.int64))

    #print(con_lcm)

    scalar_factors_construction=con_lcm*np.reciprocal(scalar_per_construction)

    #print(np.reciprocal(scalar_per_construction))

    #print(scalar_factors_construction)

    #Change this if you want to change per how many pops you optimize for
    per_labor = 100

    #Calculating the scalar factor for outputs and inputs when optimizing for labor
    scalar_per_labor=df['Labor'].values/per_labor

    #print(df['Labor'].values)

    #print(scalar_per_labor)

    #Least common multiplier to make everything into integers
    lab_lcm = np.lcm.reduce(scalar_per_labor.astype(np.int64))

    #print(lab_lcm)

    scalar_factors_labor=lab_lcm*np.reciprocal(scalar_per_labor)

    #print(scalar_factors_labor)

    #Separating the inputs and outputs of all buildings included
    df_inp=df.filter(like='Inp', axis=1)
    df_out=df.filter(like='Out', axis=1)
    #Taking into account potential throughput bonuses
    t_bonuses=1+df['TBonus'].values


    # In[107]:


    #Code for the function which preps all the math so that we can perform optimization!

    def optimization_function(scalar_factors, inp, out, tbonus):
        #Scale the inputs and outputs based on the correct scalar dependent on the maximum consturction in data set
        #Fancy way of saying that because we need to add and subtract different goods' outputs and inputs from one-another
        #And we need to have all factors be whole numbers, we have to multiply all outputs and inputs such that this is possible!
        #Otherwise we couldn't make constraits like logging camps eq.eff. is equal to eq.eff. of steel mills (they have different construction)
        inp=inp.mul(scalar_factors, axis=0).astype(int)
        out=out.mul(scalar_factors, axis=0).astype(int)
        inp=inp.mul(tbonus, axis=0).astype(int)
        out=out.mul(tbonus, axis=0).astype(int)
        
        #print(df_out)
        #print(df_inp)
        
        #Combine inputs and outputs into one equivalent net dataframe
        df_eq_net=out.sub(inp.values)
        
        #print(df_eq_net)

        #Creating our function which we will minimize! 
        #Note: since we actually want to maximize but linear optimization (SciPy) is what it is, we just reverse the function and minimize instead!
        
        df_obj_func=df_eq_net.cumsum()*-1
        
        #print(df_obj_func)
        
        c=df_obj_func.iloc[len(tbonus)-1].values
        
        #print(c)

        #Creating the lhs adn rhs matrices used in our linear optimization problem (also the whole reason this program has been written)
        #As long as the eq.eff. are within 1 of each other, the results are valid enough so we use upperbounds and not equalities
        #Technically it would be better to use equalities, but after ~8 buildings, it becomes infeasible to solve with linear optimization
        
        A=[]
        
        rhs=[]
        
        #for i in range(len(scalar_factors)):
        #    for j in range(i+1,len(scalar_factors)):
        #        A.append(df_eq_net.iloc[i].values-df_eq_net.iloc[j].values)
        #        rhs.append(1)
        #        A.append(-df_eq_net.iloc[i].values+df_eq_net.iloc[j].values)
        #        rhs.append(1)
        
        
        #Testing results with fewer constraints
        #As we want the eq.eff. to be at most 1 from each other, we create to constraints per equation:
        #One for <1 and one for >-1. However we can only use <, thus we multiply equation by -1 and get the equivalent equation!
        for i in range(len(scalar_factors)):
            if i == len(scalar_factors)-1:
                A.append(df_eq_net.iloc[i].values-df_eq_net.iloc[0].values)
                rhs.append(1)
                A.append(-df_eq_net.iloc[i].values+df_eq_net.iloc[0].values)
                rhs.append(1)
            else:
                A.append(df_eq_net.iloc[i].values-df_eq_net.iloc[i+1].values)
                rhs.append(1)
                A.append(-df_eq_net.iloc[i].values+df_eq_net.iloc[i+1].values)
                rhs.append(1)
        
        return c, A, rhs


    # In[108]:


    #Calling the function for both construction and labor! Feel free to comment the other out if you're not insterested in the results

    c_con, A_con, rhs_con = optimization_function(scalar_factors_construction, df_inp, df_out, t_bonuses)

    c_lab, A_lab, rhs_lab = optimization_function(scalar_factors_labor, df_inp, df_out, t_bonuses)


    # In[109]:


    #Creating the bounds for each good
    #Aka set it so that we don't get steel prices at like 1000+, but rather keep the prices in the +-75% range
    #These are the same regardless of whether we optimize for labor or for construction

    boundaries=[]

    for bp in BASEPRICES:
        boundaries.append((bp*0.25, bp*1.75))

    #print(boundaries)


    # In[110]:


    #Linear optimization if it's possible!

    #print(c_con)

    res_con=linprog(c_con, A_ub=A_con, b_ub=rhs_con, bounds=boundaries)
    #print(res_con.success)
    #print(res_con)
    con_result=res_con.x

    res_lab=linprog(c_lab, A_ub=A_lab, b_ub=rhs_lab, bounds=boundaries)
    #print(res_lab)
    #print(res_lab)
    lab_result=res_lab.x



    # In[111]:


    #Transforming a row in an A-matrix to be usable by mystic as constraints
    #Deprecated as using all these constraints with mystic bricks normal PCs...

    #def transformARow(x):
    #    first=0
    #
        #Find the first non-zero variable
    #    for i in range(len(x)):
    #        if x[i]!=0:
    #            first=i
    #            break

        #Solve for the first non-zero variable
    #    x=-1*x/x[first]
    #    x[first]=x[first]*-1

        #Return index of the variable solved for and the equation
    #    return first, x


    # In[112]:


    #Constraints for mystic - This is all deprecated as the amount of constraints bricks the solver on normal PCs


    #and_ = mystic.constraints.and_

    #eqns_con_string = ''

    #for co in A_con:
    #    rhs = ''
    #    first, vector = transformARow(co)
    #    lhs = str(x[first])+' = '
    #    for i in range(first+1,len(x)):
    #        rhs += vector[i]*x[i]
    #    eqns_con_string+= str(lhs)+str(rhs)+'\n'


    #Adding boundaries to constraints

    #for i in range(len(boundaries)):
    #    lhs='x'+str(i)
    #    eqns_con_string+=lhs+' >= '+str(boundaries[i][0])+'\n'
    #    eqns_con_string+=lhs+' <= '+str(boundaries[i][1])+'\n'

    #Removing unnecessary '_' characters so that mystic accepts the input
    #eqns_con_string = eqns_con_string.replace("_", "")

    #print(eqns_con_string)


    #cons_con = ms.generate_constraint(ms.generate_solvers(eqns_con_string), join=and_)



    # In[113]:


    #Initializing problem and variables (this is from an old try at PULP... Just so happens that the dictionary is useful!)
    prob = LpProblem("construction_problem", LpMinimize)
    x = pulp.LpVariable.dicts("x", range(len(BASEPRICES)), cat="Continuous")


    #Switching to mystic....
    con_mon = VerboseMonitor(10)

    #Objective function which we want to minimize
    def con_objective(x):
        return np.dot(c_con, x)


    #Penalty function for mystic. Essentially tries to make all eq.eff. scores to be equal, but not exactly since it's impossible on normal PCs

    #As we use absolute values in the penalty function, we can get rid of half of the constraints!
    A_con_penalty = []
    for i in range(0, len(A_con), 2):
        A_con_penalty.append(A_con[i])

    def con_penalty(x):
        max=0;
        #Loop through all constraints and find out the maximum deviation amongst eq.eff. scores
        for co in A_con_penalty:
            a=abs(np.dot(co,x))
            
            if a > max:
                max = a
        #Return the maximum difference in all eq.eff. scores found
        return max

    @mystic.penalty.quadratic_equality(con_penalty)
    def mystic_penalty_con(x):
        return 0.0

    #If linear optimization not possible, try global optimization with diffev2

    if not res_con.success:
        #If linear optimization not possible, then revert multiplying by lcm as we don't need to have everything be integers anymore
        A_con_penalty=np.divide(A_con_penalty, con_lcm)
        #And because we formed the objective function by adding together things multiplied by lcm, we need to also divide an additional amount
        c_con=np.divide(c_con, con_lcm*len(t_bonuses))
        #Trying global optimization
        con_result = diffev2(con_objective, x0=con_try, bounds=boundaries, penalty=mystic_penalty_con, itermon=con_mon, npop=500, maxiter=MAX_NUMBER_ITERATIONS, ftol=1e-8)

    #If global optimization is not possible for one reason or another, try and brute force the best possible solution in a reasonable time frame

    k=False

    if(len(con_result)==len(con_try)):
        if(np.array_equal(con_result, con_try)):
            print("Brute force")
            mon = VerboseMonitor(10)
            con_result = fmin(con_objective, x0=con_try, bounds=boundaries, penalty=mystic_penalty_con, itermon=mon, npop=500, maxf=MAX_NUMBER_ITERATIONS*5, maxiter=MAX_NUMBER_ITERATIONS*10, ftol=1e-7)
            k=True
    else:
        #If mystic bugs out and tries to pass some arbitrary values as optimal
        if(con_result[3]==0):
            print("Brute force")
            mon = VerboseMonitor(10)
            con_result = fmin(con_objective, x0=con_try, bounds=boundaries, penalty=mystic_penalty_con, itermon=mon, npop=500, maxf=MAX_NUMBER_ITERATIONS*5, maxiter=MAX_NUMBER_ITERATIONS*10, ftol=1e-7)
            k=True


    # In[114]:


    #Same calcs for labor


    lab_mon = VerboseMonitor(10)

    #Objective function which we want to minimize
    def lab_objective(x):
        return np.dot(c_lab, x)


    #Penalty function for mystic. Essentially tries to make all eq.eff. scores to be equal, but not exactly since it's impossible on normal PCs

    #As we use absolute values in the penalty function, we can get rid of half of the constraints!
    A_lab_penalty = []
    for i in range(0, len(A_lab), 2):
        A_lab_penalty.append(A_lab[i])

    def lab_penalty(x):
        max=0;
        #Loop through all constraints and find out the maximum deviation amongst eq.eff. scores
        for co in A_lab_penalty:
            a=abs(np.dot(co,x))  
            if a > max:
                max = a
        #Return the maximum difference in all eq.eff. scores found
        return max

    @mystic.penalty.quadratic_equality(lab_penalty)
    def mystic_penalty_lab(x):
        return 0.0

    #If linear optimization not possible, try global optimization with diffev2

    if not res_lab.success:
        #If linear optimization not possible, then revert multiplying by lcm as we don't need to have everything be integers anymore
        A_lab_penalty=np.divide(A_lab_penalty, lab_lcm)
        #And because we formed the objective function by adding together things multiplied by lcm, we need to also divide an additional amount
        c_lab=np.divide(c_lab, lab_lcm*len(t_bonuses))
        #Trying global optimization
        lab_result = diffev2(lab_objective, x0=lab_try, bounds=boundaries, penalty=mystic_penalty_lab, itermon=lab_mon, npop=500, maxiter=MAX_NUMBER_ITERATIONS, ftol=1e-8)

    #If global optimization is not possible for one reason or another, try and brute force the best possible solution in a reasonable time frame

    j=False

    if(len(lab_result)==len(lab_try)):
        if(np.array_equal(lab_result, lab_try)):
            print("Brute force")
            mon = VerboseMonitor(10)
            lab_result = fmin(lab_objective, x0=lab_try, bounds=boundaries, penalty=mystic_penalty_lab, itermon=mon, npop=500, maxf=MAX_NUMBER_ITERATIONS*5, maxiter=MAX_NUMBER_ITERATIONS*10, ftol=1e-7)
            j=True
    else:
        #If mystic bugs out and tries to pass some arbitrary values as optimal
        if(lab_result[3]==0):
            print("Brute force")
            mon = VerboseMonitor(10)
            lab_result = fmin(lab_objective, x0=lab_try, bounds=boundaries, penalty=mystic_penalty_lab, itermon=mon, npop=500, maxf=MAX_NUMBER_ITERATIONS*5, maxiter=MAX_NUMBER_ITERATIONS*10, ftol=1e-7)
            j=True


    # In[115]:


    #Making the optimal prices per construction more readable

    procentages=[]

    for i in range(len(con_result)):
        procentages.append(str(round(((con_result[i]/BASEPRICES[i])-1)*100, 1))+"%")

    d = {'Good':GOODSNAMES, 'Base price':BASEPRICES, 'Optimal price per construction':con_result, 'Procentage':procentages}
    readable_df=pd.DataFrame(data=d)

    if not res_con.success:
        if k:
            print("Optimization was unfortunately not possible and thus these prices are simply 'guesses' that tend to the right direction")
        else:
            print("Note that linear optimization could not be performed and these prices are simply the best your PC and this program could calculate")

    readable_df.sort_values('Good')


    # In[116]:


    #Making the optimal prices per labor more readable

    procentages=[]

    if len(lab_result)!=len(BASEPRICES):
        lab_result=lab_result[0].tolist()

    for i in range(len(lab_result)):
        procentages.append(str(round(((lab_result[i]/BASEPRICES[i])-1)*100, 1))+"%")

    d = {'Good':GOODSNAMES, 'Base price':BASEPRICES, 'Optimal price per labor':lab_result, 'Procentage':procentages}
    readable_df2=pd.DataFrame(data=d)

    if not res_lab.success:
        if j:
            print("Optimization was unfortunately not possible and thus these prices are simply 'guesses' that tend to the right direction")
        else:
            print("Note that linear optimization could not be performed and these prices are simply the best your PC and this program could calculate")
        
    readable_df2.sort_values('Good')


    # In[117]:


    #Writing the results into an Excel-sheet for the purposes of an .exe

    with pd.ExcelWriter('OptimizedPrices.xlsx') as writer:
        readable_df.to_excel(writer, sheet_name='Optimized for construction')
        readable_df2.to_excel(writer, sheet_name='Optimized for labor')

    #Marking goods that aren't included in the calculations
    workbook = load_workbook(filename='OptimizedPrices.xlsx')
    ws4 = workbook['Optimized for construction']
    for i in range(len(c_con)):
        if c_con[i] == 0:
            ws4.cell(row = i+2, column = 7).value = 'Does not affect net value added'
    workbook.save('OptimizedPrices.xlsx')

    workbook = load_workbook(filename='OptimizedPrices.xlsx')
    ws4 = workbook['Optimized for labor']
    for i in range(len(c_lab)):
        if c_lab[i] == 0:
            ws4.cell(row = i+2, column = 7).value = 'Does not affect net value added'
    workbook.save('OptimizedPrices.xlsx')


    #Also writing 'warnings' to the Excel-sheet

    if not res_con.success:
        workbook = load_workbook(filename='OptimizedPrices.xlsx')
        ws4 = workbook['Optimized for construction']
        #Adding a message to encourage running the program again!
        if abs(con_objective(con_try)-con_objective(con_result)) > 0.005 or abs(con_penalty(con_try)-con_penalty(con_result)) > 0.005:
                ws4.cell(row = 4, column = 9).value = "I recommend running the program again without touching the values in this sheet as further optimization can be done!"
        else:
            ws4.cell(row = 4, column = 9).value = "Further optimization might be possible, but would not yield significant results"
        #If solution used brute force
        if k:
            ws4.cell(row = 2, column = 9).value = "Further optimization was unfortunately not possible and thus these prices are the best 'guesses' reasonably possible"
        #If solution used diffenv2
        else:
            ws4.cell(row = 2, column = 9).value = "Note that linear optimization could not be performed and these prices are simply the best your PC and this program could calculate in this round"

        workbook.save('OptimizedPrices.xlsx')

    if not res_lab.success:
        workbook = load_workbook(filename='OptimizedPrices.xlsx')
        ws4 = workbook['Optimized for labor']
        #Addint a message to encourage running the program again!
        if abs(lab_objective(lab_try)-lab_objective(lab_result)) > 0.005 or abs(lab_penalty(lab_try)-lab_penalty(lab_result)) > 0.005:
                ws4.cell(row = 4, column = 9).value = "I recommend running the program again without touching the values in this sheet as further optimization can be done!"
        else:
            ws4.cell(row = 4, column = 9).value = "Further optimization might be possible, but would not yield significant results"
        #If solution used brute force
        if j:
            ws4.cell(row = 2, column = 9).value = "Further optimization was unfortunately not possible and thus these prices are the best 'guesses' reasonably possible"
        #If solution used diffenv2
        else:
            ws4.cell(row = 2, column = 9).value = "Note that linear optimization could not be performed and these prices are simply the best your PC and this program could calculate in this round"
        workbook.save('OptimizedPrices.xlsx')



    # In[118]:


    #print("Program has finished!")
    statusText.set("Program has finished!")
    app.update()


    # In[119]:


    #Optional funsies for those testing with the notebook!

    #print(" Construction stuff. The smaller the result and penalty, the better")
    #print(con_objective(con_result))
    #print(con_penalty(con_result))

    #print("\n Labor stuff. The smaller the result and penalty, the better")
    #print(lab_objective(lab_result))
    #print(lab_penalty(lab_result))


    #print("Improvement compared to initial guess: ")

    #print("Construction objective: ")
    cobj=con_objective(con_try)-con_objective(con_result)
    if con_objective(con_try)!=0:
    #print(cobj)
        #print("Construction objective improvement of about: "+str(round(-(cobj/con_objective(con_try))*100,4))+"%")
        currentText = statusText.get()
        statusText.set(currentText+'\n'+"Construction objective improvement of about: "+str(round(-(cobj/con_objective(con_try))*100,4))+"%")
        app.update()
    #print("Construction penalty: ")
    copen=con_penalty(con_try)-con_penalty(con_result)
    if con_penalty(con_try)!=0:
    #print(copen)
        #print("Construction penalty improvement of about: "+str(round((copen/con_penalty(con_try))*100,4))+"%")
        currentText = statusText.get()
        statusText.set(currentText+'\n'+"Construction penalty improvement of about: "+str(round((copen/con_penalty(con_try))*100,4))+"%")
        app.update()
    #print("Labor objective: ")
    lobj=lab_objective(lab_try)-lab_objective(lab_result)
    if lab_objective(lab_try)!=0:
    #print(lobj)
        #print("Labor objective improvement of about: "+str(round(-(lobj/lab_objective(lab_try))*100,4))+"%")
        currentText = statusText.get()
        statusText.set(currentText+'\n'+"Labor objective improvement of about: "+str(round(-(lobj/lab_objective(lab_try))*100,4))+"%")
        app.update()
    #print("Labor penalty: ")
    lpen=lab_penalty(lab_try)-lab_penalty(lab_result)
    if lab_penalty(lab_try)!=0:
    #print(lpen)
        #print("Labor penalty improvement of about: "+str(round((lpen/lab_penalty(lab_try))*100,4))+"%")
        currentText = statusText.get()
        statusText.set(currentText+'\n'+"Labor penalty improvement of about: "+str(round((lpen/lab_penalty(lab_try))*100,4))+"%")
        app.update()

# In[ ]:




